"""
MMG Pricing Sync Script
=======================
Run this script whenever you update the Excel pricing file.
It reads the Excel, updates the JSON master, and regenerates the readable .md.

Usage (from anywhere, just double-click or run in terminal):
    python scripts/sync-pricing.py

The Excel lives in Google Drive and auto-syncs to your phone.
Edit it on phone or PC, then run this script on PC to push to the app.

What it does:
  1. Reads MMG_COURSE_PRICING_MASTER.xlsx (Google Drive Reference folder)
  2. Updates the internal block in MMG_COURSE_PRICING_MASTER.json
  3. Rewrites MMG_COURSE_PRICING_MASTER_READABLE.md

Fields managed by this script (from Excel):
  - low, mid, peak green fees
  - buggyMin, buggyMax
  - clubsMin, clubsMax
  - toDiscount (TO discount %, null if blank)
  - gcType, golfcard, licence

All other JSON fields (calculator, dayCost, research) are NOT touched.
"""

import json
import os
import sys
from datetime import date
from pathlib import Path

# ── Paths ────────────────────────────────────────────────────────────────────
DRIVE_REF = Path(r"C:\Users\andyg\My Drive\Mr Mallorca Golf\Reference")
JSON_PATH = DRIVE_REF / "MMG_COURSE_PRICING_MASTER_DO-NOT-EDIT.json"
EXCEL_PATH = DRIVE_REF / "MMG_COURSE_PRICING_MASTER_EDIT-THIS.xlsx"
READABLE_PATH = DRIVE_REF / "MMG_COURSE_PRICING_MASTER_READABLE_DO-NOT-EDIT.md"

# ── Helpers ──────────────────────────────────────────────────────────────────
def val(v):
    """Return None if blank/NaN, else the value."""
    if v is None:
        return None
    try:
        import math
        if math.isnan(float(v)):
            return None
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return None if s in ("", "nan", "None", "NaN") else v

def to_int(v):
    v = val(v)
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def to_bool(v):
    v = val(v)
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("true", "yes", "1", "y")

def fmt_eur(v):
    if v is None:
        return "-"
    return f"EUR {int(v)}"

def fmt_range(lo, hi):
    if lo is None and hi is None:
        return "-"
    if lo == hi or hi is None:
        return fmt_eur(lo)
    if lo is None:
        return fmt_eur(hi)
    return f"EUR {int(lo)}-{int(hi)}"

# ── Step 1: Read Excel ────────────────────────────────────────────────────────
def read_excel():
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
        import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        course_id = str(d.get("courseId", "")).strip()
        if course_id:
            rows[course_id] = d
    return rows

# ── Step 2: Update JSON ───────────────────────────────────────────────────────
def update_json(excel_rows):
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    courses = data["appPricing"]["courses"]
    changed = []

    for course_id, d in excel_rows.items():
        if course_id not in courses:
            print(f"  WARNING: {course_id} not found in JSON, skipping")
            continue

        internal = courses[course_id].setdefault("internal", {})
        old = dict(internal)

        internal["low"]        = to_int(d.get("low"))
        internal["mid"]        = to_int(d.get("mid"))
        internal["peak"]       = to_int(d.get("peak"))
        internal["buggyMin"]   = to_int(d.get("buggyMin")) or 0
        internal["buggyMax"]   = to_int(d.get("buggyMax")) or 0
        internal["clubsMin"]   = to_int(d.get("clubsMin")) or 0
        internal["clubsMax"]   = to_int(d.get("clubsMax")) or 0
        internal["toDiscount"] = to_int(d.get("toDiscount"))
        internal["gcType"]     = val(d.get("gcType")) or "none"
        internal["golfcard"]   = to_bool(d.get("golfcard"))
        internal["licence"]    = to_bool(d.get("licence"))

        if internal != old:
            changed.append(courses[course_id]["name"])

    data["updated"] = str(date.today())

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return data, changed

# ── Step 3: Regenerate readable .md ──────────────────────────────────────────
def regen_readable(data):
    courses = data["appPricing"]["courses"]
    research = data["research"]["courses"]

    lines = [
        "# MMG Course Pricing Master",
        "",
        f"Generated from: `{JSON_PATH}`",
        f"Generated: {date.today()}",
        "",
        "Edit the Excel (`MMG_COURSE_PRICING_MASTER.xlsx`) for changes, then run `python scripts/sync-pricing.py` to sync.",
        "",
        "## How To Read This",
        "",
        "- `App` rows are the exact values synced into internal, day-cost, and calculator.",
        "- `Research` rows are the public-market evidence and can contain exact values, ranges, or notes.",
        "- `TO Discount` is the tour operator green fee discount (%) for that course.",
        "- Ranges such as `EUR 35-48` usually mean season, time, booking channel, or other conditions change the price.",
        "",
    ]

    # Sort alphabetically by display name
    sorted_keys = sorted(courses.keys(), key=lambda k: courses[k].get("name", k))

    for key in sorted_keys:
        c = courses[key]
        name = c.get("name", key)
        internal = c.get("internal", {})
        r = research.get(key, {})
        prices = r.get("prices", {})

        def rp(field):
            p = prices.get(field)
            if not p:
                return "-"
            if isinstance(p, dict):
                t = p.get("type")
                if t == "exact":
                    return f"EUR {p['amount']}"
                elif t == "range":
                    return f"EUR {p['min']}-{p['max']}"
                elif t in ("note", "free"):
                    return str(p.get("raw", "-"))
            return "-"

        to_disc = internal.get("toDiscount")
        to_str = f"{to_disc}%" if to_disc else "-"

        lines += [
            f"## {name}",
            "",
            "| Type | Value 1 | Value 2 | Value 3 |",
            "|---|---:|---:|---|",
            f"| App green fees | Low {fmt_eur(internal.get('low'))} | Mid {fmt_eur(internal.get('mid'))} | Peak {fmt_eur(internal.get('peak'))} |",
            f"| App buggy | {fmt_range(internal.get('buggyMin'), internal.get('buggyMax'))} | GC type: {internal.get('gcType','?')} | Golf card: {'Yes' if internal.get('golfcard') else 'No'} |",
            f"| App clubs | {fmt_range(internal.get('clubsMin'), internal.get('clubsMax'))} | Licence: {'Yes' if internal.get('licence') else 'No'} | TO discount: {to_str} |",
            f"| Research status | {r.get('certainty','-')} | Publish safe: {r.get('publishSafe','-')} | {r.get('pricingStatus','-')} |",
            f"| Research green fees | Low {rp('Adult 18 Low EUR')} | Peak {rp('Adult 18 Peak EUR')} | Twilight {rp('Twilight EUR')} |",
            f"| Research extras | Buggy {rp('Buggy EUR')} | Clubs {rp('Club Hire EUR')} | Range {rp('Range EUR')} |",
            f"| Season notes | {r.get('seasonNotes','-')} |  |  |",
            f"| Source notes | {r.get('sourceNotes','-')} |  |  |",
            f"| Confidence gap | {r.get('whyNot100PercentConfirmed','-') or '-'} |  |  |",
            "",
        ]

    with open(READABLE_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

# ── Step 4: Generate Excel (run once to create, or to reset) ─────────────────
def generate_excel(data):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        os.system(f'"{sys.executable}" -m pip install openpyxl --quiet')
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    courses = data["appPricing"]["courses"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"

    headers = [
        "courseId", "name",
        "low", "mid", "peak",
        "buggyMin", "buggyMax",
        "clubsMin", "clubsMax",
        "toDiscount",
        "gcType", "golfcard", "licence",
    ]

    header_fill = PatternFill("solid", fgColor="1A3A2A")
    arabella_fill = PatternFill("solid", fgColor="FFF3CD")
    header_font = Font(bold=True, color="FFFFFF")

    # Write header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    arabella_ids = {"son_quint", "son_muntaner", "son_vida", "palma_pp"}

    sorted_keys = sorted(courses.keys(), key=lambda k: courses[k].get("name", k))

    for row_i, key in enumerate(sorted_keys, 2):
        c = courses[key]
        internal = c.get("internal", {})
        row_data = [
            key,
            c.get("name", key),
            internal.get("low"),
            internal.get("mid"),
            internal.get("peak"),
            internal.get("buggyMin"),
            internal.get("buggyMax"),
            internal.get("clubsMin"),
            internal.get("clubsMax"),
            internal.get("toDiscount"),
            internal.get("gcType", "none"),
            internal.get("golfcard", False),
            internal.get("licence", False),
        ]
        for col, val_cell in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val_cell)
            if key in arabella_ids:
                cell.fill = arabella_fill

    # Column widths
    widths = [16, 22, 7, 7, 7, 9, 9, 9, 9, 11, 10, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a note sheet
    ws2 = wb.create_sheet("HOW TO USE")
    notes = [
        ["MMG Course Pricing — How to Edit"],
        [""],
        ["1. Edit any value in the Pricing sheet"],
        ["   Yellow rows = Arabella courses (Son Quint, Son Muntaner, Son Vida, Palma P+P)"],
        [""],
        ["2. Save the file (Ctrl+S or auto-saves in Google Sheets on phone)"],
        [""],
        ["3. On PC, run:   python scripts/sync-pricing.py"],
        ["   This updates the JSON master + rebuilds the readable .md"],
        ["   The internal.mrmallorcagolf.com app is then updated next time sync-pricing.js runs."],
        [""],
        ["Key columns:"],
        ["  low / mid / peak     = green fee tiers (EUR)"],
        ["  buggyMin / buggyMax  = buggy hire range (EUR, 0 if included or N/A)"],
        ["  clubsMin / clubsMax  = club hire range (EUR)"],
        ["  toDiscount           = tour operator discount % (leave blank if none)"],
        ["  gcType               = 2for1 / disc / special / pkg / none"],
        ["  golfcard             = TRUE if Golf Card Mallorca accepted"],
        ["  licence              = TRUE if 3 EUR daily golf licence required"],
        [""],
        ["DO NOT change courseId column — it links to the JSON."],
    ]
    for r, row in enumerate(notes, 1):
        ws2.cell(row=r, column=1, value=row[0] if row else "")
    ws2.column_dimensions["A"].width = 70
    ws2["A1"].font = Font(bold=True, size=13)

    wb.save(EXCEL_PATH)
    print(f"  Excel written: {EXCEL_PATH}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("MMG Pricing Sync")
    print("=" * 40)

    # If Excel doesn't exist yet, generate it from JSON first
    if not EXCEL_PATH.exists():
        print("Excel not found — generating from JSON master...")
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        generate_excel(data)
        print("Done. Edit the Excel, then run this script again to sync back.")
        return

    print(f"Reading Excel: {EXCEL_PATH.name}")
    excel_rows = read_excel()
    print(f"  Found {len(excel_rows)} courses in Excel")

    print("Updating JSON...")
    data, changed = update_json(excel_rows)
    if changed:
        print(f"  Changed: {', '.join(changed)}")
    else:
        print("  No pricing changes detected")

    print("Regenerating readable .md...")
    regen_readable(data)
    print(f"  Written: {READABLE_PATH.name}")

    print()
    print("All done. Changes are live in the JSON.")
    print("Run  node scripts/sync-pricing.js  to push to the internal app.")

if __name__ == "__main__":
    main()
